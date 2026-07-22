"""Benchmark xlsx-creation-editing operations against the Apache POI corpus.

Design:
- Uses POI_PATH env var (default: poi/test-data/spreadsheet relative to repo root)
  per the cross-platform convention; all paths via pathlib.Path.
- tempfile.mkdtemp() for all output to avoid hardcoded paths.
- Measures four wall-clock operations per file: open, iterate, write_100, validate.
- Prints a fixed-width table; accepts --quick for first 3 fixtures only.
"""

from __future__ import annotations

import argparse
import os
import sys
import time
from pathlib import Path
from tempfile import TemporaryDirectory

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))


def _default_poi_path() -> tuple[Path, str]:
    baseline_relative_path = Path("test-data/spreadsheet")
    workspace_root = ROOT.parents[1]
    return workspace_root / "poi" / baseline_relative_path, str(baseline_relative_path)


def _deps() -> tuple[object, object, object, object, object]:
    import openpyxl

    from scripts.create_xlsx import create_workbook, save, write_cell
    from scripts.validate_xlsx import validate_xlsx

    return openpyxl, create_workbook, save, write_cell, validate_xlsx

BENCH_FIXTURES = [
    "sample.xlsx",
    "simple-monthly-budget.xlsx",
    "ExcelTables.xlsx",
    "Formatting.xlsx",
    "ConditionalFormattingSamples.xlsx",
    "123233_charts.xlsx",
    "styles.xlsx",
    "DateFormatTests.xlsx",
    "AverageTaxRates.xlsx",
    "simple-table-named-range.xlsx",
]


def _time_open(path: Path) -> float:
    openpyxl, _, _, _, _ = _deps()
    start = time.perf_counter()
    openpyxl.load_workbook(str(path), data_only=True)
    return (time.perf_counter() - start) * 1000


def _time_iterate(path: Path) -> float:
    openpyxl, _, _, _, _ = _deps()
    start = time.perf_counter()
    wb = openpyxl.load_workbook(str(path), data_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for _ in row:
                pass
    return (time.perf_counter() - start) * 1000


def _time_write_100() -> float:
    _, create_workbook, save, write_cell, _ = _deps()
    with TemporaryDirectory() as tmpdir:
        start = time.perf_counter()
        wb = create_workbook()
        ws = wb.active
        for r in range(1, 11):
            for c in range(1, 11):
                write_cell(ws, r, c, r * c)
        save(wb, Path(tmpdir) / "bench_write.xlsx")
        return (time.perf_counter() - start) * 1000


def _time_validate(path: Path) -> float:
    _, _, _, _, validate_xlsx = _deps()
    start = time.perf_counter()
    validate_xlsx(path)
    return (time.perf_counter() - start) * 1000


def main() -> None:
    parser = argparse.ArgumentParser(description="Benchmark xlsx-creation-editing")
    parser.add_argument("--quick", action="store_true", help="First 3 fixtures only")
    parser.add_argument(
        "--poi-path",
        type=Path,
        default=None,
        help="Optional path to Apache POI test-data/spreadsheet",
    )
    args = parser.parse_args()

    default_poi_path, baseline_hint = _default_poi_path()

    poi_path = Path(
        os.environ.get(
            "POI_PATH",
            str(args.poi_path or default_poi_path),
        )
    )

    if not poi_path.exists():
        print(
            "ERROR: spreadsheet fixture path not found: "
            f"{poi_path}. Expected Apache POI baseline under {baseline_hint}."
        )
        sys.exit(1)

    fixtures = BENCH_FIXTURES[:3] if args.quick else BENCH_FIXTURES
    col_w = max(len(f) for f in fixtures) + 2

    header = f"{'Fixture':<{col_w}}  {'open':>8}  {'iterate':>8}  {'write_100':>9}  {'validate':>9}"
    print(header)
    print("-" * len(header))

    for fname in fixtures:
        path = poi_path / fname
        if not path.exists():
            print(f"{fname:<{col_w}}  (not found)")
            continue
        try:
            t_open = _time_open(path)
            t_iter = _time_iterate(path)
            t_write = _time_write_100()
            t_val = _time_validate(path)
            print(
                f"{fname:<{col_w}}  {t_open:7.1f}ms  {t_iter:7.1f}ms"
                f"  {t_write:8.1f}ms  {t_val:8.1f}ms"
            )
        except Exception as exc:
            print(f"{fname:<{col_w}}  ERROR: {exc}")


if __name__ == "__main__":
    main()
