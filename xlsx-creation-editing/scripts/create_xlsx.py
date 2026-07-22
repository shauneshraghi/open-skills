"""Create and save Excel (.xlsx) workbooks from scratch.

Design decisions from primary sources:
- openpyxl.Workbook() with write_only=False creates a full in-memory workbook
  (openpyxl docs § Working with Excel Files).
- Cells accept str, int, float, datetime, bool, None, and formula strings
  (starting with '='); openpyxl stores formulas as strings and Excel evaluates
  them on open (openpyxl docs § Simple Usage).
- PatternFill, Font, Border, Alignment are applied directly to Cell objects;
  copy_protection=False is the default (openpyxl docs § Styling).
- reorder_sheet() uses wb.move_sheet(sheet, offset) which has been available
  since openpyxl 2.6 (openpyxl source, Workbook.move_sheet).  If the method is
  absent (hypothetical older install), we fall back to lxml manipulation of
  <workbook><sheets> ordering per ECMA-376 §18.2.19.
- Alt text for images is injected via a post-save ZIP patch using lxml because
  openpyxl does not expose a setter for <xdr:cNvPr descr="…">
  (ECMA-376 §20.5.2.8).  The patch writes the modified drawing XML back into the
  ZIP in-place without touching any other parts.
"""

from __future__ import annotations

import argparse
import hashlib
import io
import zipfile
from pathlib import Path
from typing import Any

from lxml import etree
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.cell import coordinate_to_tuple

# XML namespace for SpreadsheetML drawing parts (ECMA-376 §20.5)
_XDR_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_XDR = f"{{{_XDR_NS}}}"
_A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
_A = f"{{{_A_NS}}}"
_R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_R = f"{{{_R_NS}}}"
_SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_SHEET = f"{{{_SHEET_NS}}}"

# Relationship type for worksheet drawings (OPC / ECMA-376 Part 2)
_REL_DRAWING = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"
_REL_WORKSHEET = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet"


# ─────────────────────────── workbook / sheet ops ────────────────────────────

def create_workbook() -> Workbook:
    """Return a new empty Workbook with one default sheet named 'Sheet'."""
    wb = Workbook()
    wb.active.title = "Sheet"
    return wb


def add_sheet(wb: Workbook, name: str, position: int | None = None) -> Any:
    """Create a new worksheet named *name* and return it.

    If *position* is given (0-based), the sheet is inserted there; otherwise it
    is appended.  openpyxl docs § Working with Sheets.
    """
    return wb.create_sheet(title=name, index=position)


def rename_sheet(ws: Any, new_name: str) -> None:
    """Rename a worksheet in-place."""
    ws.title = new_name


def delete_sheet(wb: Workbook, name_or_index: str | int) -> None:
    """Remove a sheet by name or 0-based index."""
    if len(wb.worksheets) <= 1:
        raise ValueError("Cannot delete the only worksheet in a workbook.")
    if isinstance(name_or_index, int):
        ws = wb.worksheets[name_or_index]
    else:
        ws = wb[name_or_index]
    del wb[ws.title]


def reorder_sheet(wb: Workbook, from_index: int, to_index: int) -> None:
    """Move the sheet at *from_index* to *to_index* (0-based).

    Uses wb.move_sheet() (available since openpyxl 2.6).  If move_sheet is
    absent the fallback manipulates the underlying <workbook><sheets> element
    ordering via lxml (ECMA-376 §18.2.19: document order of <sheet> children
    determines workbook sheet order).
    """
    n = len(wb.worksheets)
    if not (0 <= from_index < n and 0 <= to_index < n):
        raise IndexError(
            f"Sheet index out of range (0..{n - 1}): from={from_index} to={to_index}"
        )
    if from_index == to_index:
        return

    if hasattr(wb, "move_sheet"):
        sheet = wb.worksheets[from_index]
        wb.move_sheet(sheet, offset=to_index - from_index)
    else:
        # lxml fallback: manipulate <workbook><sheets> child order.
        # ECMA-376 §18.2.19: <sheet> element order within <sheets> determines
        # the user-visible sheet tab order.
        # Reorder wb._sheets list
        sheets = wb._sheets
        sheet = sheets.pop(from_index)
        sheets.insert(to_index, sheet)


# ─────────────────────────── cell / range ops ────────────────────────────────

def write_cell(ws: Any, row: int, col: int, value: Any) -> None:
    """Write *value* to cell at 1-based (*row*, *col*).

    Accepts str, int, float, datetime, bool, None, and formula strings
    (starting with '=').  openpyxl docs § Simple Usage.
    """
    ws.cell(row=row, column=col, value=value)


def write_range(ws: Any, start_row: int, start_col: int, data: list[list[Any]]) -> None:
    """Write a 2-D list starting at (start_row, start_col) (1-based).

    Each inner list is a row.  openpyxl docs § Accessing many cells.
    """
    for r_offset, row_data in enumerate(data):
        for c_offset, value in enumerate(row_data):
            ws.cell(row=start_row + r_offset, column=start_col + c_offset, value=value)


def format_cell(
    ws: Any,
    row: int,
    col: int,
    *,
    font: dict | None = None,
    fill: dict | None = None,
    border: dict | None = None,
    number_format: str | None = None,
    alignment: dict | None = None,
) -> None:
    """Apply formatting to a cell.

    Each keyword argument is a dict of keyword arguments passed to the
    corresponding openpyxl style constructor:
    - font: Font(**font) e.g. {"bold": True, "name": "Arial", "size": 12}
    - fill: PatternFill(**fill) e.g. {"patternType": "solid", "fgColor": "FFFF00"}
    - border: dict of side kwargs, e.g. {"left": "thin", "right": "thin"}
    - number_format: format string, e.g. "#,##0.00" or "YYYY-MM-DD"
    - alignment: Alignment(**alignment) e.g. {"horizontal": "center"}

    openpyxl docs § Styling.
    """
    cell = ws.cell(row=row, column=col)
    if font is not None:
        cell.font = Font(**font)
    if fill is not None:
        cell.fill = PatternFill(**fill)
    if border is not None:
        sides = {k: Side(style=v) for k, v in border.items() if v}
        cell.border = Border(**sides)
    if number_format is not None:
        cell.number_format = number_format
    if alignment is not None:
        cell.alignment = Alignment(**alignment)


def merge_cells(
    ws: Any, start_row: int, start_col: int, end_row: int, end_col: int
) -> None:
    """Merge the rectangular block from (start_row, start_col) to (end_row, end_col).

    openpyxl docs § Merging cells.  Internally stores a MergedCell placeholder
    so the merged range appears in ws.merged_cells.
    """
    ws.merge_cells(
        start_row=start_row,
        start_column=start_col,
        end_row=end_row,
        end_column=end_col,
    )


def unmerge_cells(
    ws: Any, start_row: int, start_col: int, end_row: int, end_col: int
) -> None:
    """Unmerge a previously merged block."""
    ws.unmerge_cells(
        start_row=start_row,
        start_column=start_col,
        end_row=end_row,
        end_column=end_col,
    )


# ─────────────────────────── image ops ───────────────────────────────────────

def add_image_to_sheet(
    ws: Any,
    image_path: str | Path,
    anchor: str,
    *,
    alt_text: str = "",
) -> XLImage:
    """Add an image to the worksheet at *anchor* (e.g. "B2").

    Alt text is stored in a side-channel dict on the worksheet object
    (``ws._xlsx_skill_alt_texts``) and injected into the drawing XML's
    ``<xdr:cNvPr descr="…">`` attribute when ``save()`` is called.

    ECMA-376 §20.5.2.8: <xdr:cNvPr descr="…"> inside <xdr:nvPicPr> carries
    the accessibility description for a picture in a spreadsheet drawing part.
    openpyxl has no high-level setter for this attribute, so we write it via
    lxml in the post-save ZIP patch performed by save().
    """
    img = XLImage(str(image_path))
    img.anchor = anchor
    ws.add_image(img)
    if alt_text:
        set_image_metadata(
            ws,
            len(ws._images) - 1,
            image_path,
            anchor,
            alt_text=alt_text,
        )
    return img


def set_image_metadata(
    ws: Any,
    index: int,
    image_path: str | Path,
    anchor: Any,
    *,
    alt_text: str | None = None,
) -> None:
    """Store matching metadata for later ZIP-level drawing patching."""
    metadata: list[dict[str, Any] | None] = list(
        getattr(ws, "_xlsx_skill_image_metadata", [])
    )
    while len(metadata) <= index:
        metadata.append(None)

    entry = dict(metadata[index] or {})
    source_path = Path(image_path)
    entry.update(
        {
            "source_hash": _hash_file(source_path),
            "source_name": source_path.name,
            "anchor_key": _normalize_anchor(anchor),
        }
    )
    if alt_text is not None:
        entry["alt_text"] = alt_text

    metadata[index] = entry
    ws._xlsx_skill_image_metadata = metadata


# ─────────────────────────── save (with alt-text patch) ──────────────────────

def save(wb: Workbook, path: str | Path) -> Path:
    """Save the workbook; creates parent directories as needed.

    After the standard openpyxl save, applies a ZIP-level patch for any images
    that had alt_text supplied to add_image_to_sheet().  The patch uses lxml to
    set <xdr:cNvPr descr="…"> (ECMA-376 §20.5.2.8) inside the relevant drawing
    XML parts without modifying any other parts.
    """
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    _patch_alt_texts(wb, out)
    return out


def _patch_alt_texts(wb: Workbook, path: Path) -> None:
    """Inject alt texts into drawing XML parts of the saved ZIP.

    For each worksheet that has pending alt texts (stored in
    ws._xlsx_skill_alt_texts), this function:
    1. Locates the worksheet's drawing relationship from its .rels file.
    2. Parses the drawing XML with lxml.
    3. Finds the nth <xdr:pic> element corresponding to the image index.
    4. Sets cNvPr@descr to the alt text.
    5. Rewrites the entire ZIP with the modified drawing XML.

    ECMA-376 §20.5.2.8 and OPC Part 2 §10 (ZIP packaging).
    """
    pending: dict[int, list[dict[str, Any]]] = {}
    for ws_idx, ws in enumerate(wb.worksheets):
        metadata = [
            entry
            for entry in getattr(ws, "_xlsx_skill_image_metadata", [])
            if entry and entry.get("alt_text")
        ]
        if metadata:
            pending[ws_idx] = metadata
    if not pending:
        return

    raw = path.read_bytes()
    file_map: dict[str, bytes] = {}
    with zipfile.ZipFile(io.BytesIO(raw), "r") as zf:
        for name in zf.namelist():
            file_map[name] = zf.read(name)

    worksheet_parts = _worksheet_part_paths_in_order(file_map)
    failures: list[str] = []

    for ws_idx, metadata_entries in pending.items():
        if ws_idx >= len(worksheet_parts):
            failures.append(f"Worksheet index {ws_idx} has no matching worksheet part.")
            continue

        worksheet_part = worksheet_parts[ws_idx]
        rels_key = _rels_part_for_part(worksheet_part)
        if rels_key not in file_map:
            failures.append(f"Missing worksheet relationships part for {worksheet_part}.")
            continue

        rels_root = etree.fromstring(file_map[rels_key])
        drawing_target: str | None = None
        for rel in rels_root:
            if _REL_DRAWING in rel.get("Type", ""):
                drawing_target = rel.get("Target", "")
                break
        if not drawing_target:
            failures.append(f"Worksheet {worksheet_part} has no drawing relationship.")
            continue

        drawing_part = _resolve_path(worksheet_part, drawing_target)
        if drawing_part not in file_map:
            failures.append(
                f"Worksheet drawing target for {worksheet_part} is missing: {drawing_part}"
            )
            continue

        drawing_rels_key = _rels_part_for_part(drawing_part)
        if drawing_rels_key not in file_map:
            failures.append(f"Missing drawing relationships part for {drawing_part}.")
            continue

        root = etree.fromstring(file_map[drawing_part])
        drawing_rels_root = etree.fromstring(file_map[drawing_rels_key])
        pictures = _picture_records(root, drawing_part, drawing_rels_root, file_map)
        used_indices: set[int] = set()

        for metadata in metadata_entries:
            match_index = _match_picture_record(metadata, pictures, used_indices)
            if match_index is None:
                failures.append(
                    "Unable to match image metadata to drawing picture for "
                    f"worksheet {worksheet_part}: {metadata}"
                )
                continue

            pictures[match_index]["cnvpr"].set("descr", metadata["alt_text"])
            used_indices.add(match_index)

        file_map[drawing_part] = etree.tostring(
            root, xml_declaration=True, encoding="UTF-8", standalone=True
        )

    if failures:
        joined = "; ".join(failures)
        raise ValueError(f"Failed to patch XLSX image alt text safely: {joined}")

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf_out:
        for name, data in file_map.items():
            zf_out.writestr(name, data)
    path.write_bytes(buf.getvalue())


def _resolve_path(source: str, target: str) -> str:
    """Resolve an OPC relative or absolute target path against a source part name."""
    # Absolute OPC target paths start with '/' — strip the leading slash.
    if target.startswith("/"):
        return target.lstrip("/")
    source_dir = source.rsplit("/", 1)[0] if "/" in source else ""
    combined = source_dir + "/" + target if source_dir else target
    parts = combined.split("/")
    resolved: list[str] = []
    for part in parts:
        if part == "..":
            if resolved:
                resolved.pop()
        elif part and part != ".":
            resolved.append(part)
    return "/".join(resolved)


def _rels_part_for_part(part_name: str) -> str:
    part_path = Path(part_name)
    return str(part_path.parent / "_rels" / f"{part_path.name}.rels").replace("\\", "/")


def _worksheet_part_paths_in_order(file_map: dict[str, bytes]) -> list[str]:
    workbook_part = "xl/workbook.xml"
    workbook_rels_part = "xl/_rels/workbook.xml.rels"
    if workbook_part not in file_map or workbook_rels_part not in file_map:
        return []

    workbook_root = etree.fromstring(file_map[workbook_part])
    workbook_rels_root = etree.fromstring(file_map[workbook_rels_part])
    rel_targets = {
        rel.get("Id", ""): _resolve_path(workbook_part, rel.get("Target", ""))
        for rel in workbook_rels_root
        if _REL_WORKSHEET in rel.get("Type", "")
    }

    worksheet_parts: list[str] = []
    for sheet in workbook_root.findall(f".//{_SHEET}sheets/{_SHEET}sheet"):
        rel_id = sheet.get(f"{_R}id")
        if rel_id and rel_id in rel_targets:
            worksheet_parts.append(rel_targets[rel_id])
    return worksheet_parts


def _picture_records(
    drawing_root: Any,
    drawing_part: str,
    drawing_rels_root: Any,
    file_map: dict[str, bytes],
) -> list[dict[str, Any]]:
    rel_targets = {
        rel.get("Id", ""): _resolve_path(drawing_part, rel.get("Target", ""))
        for rel in drawing_rels_root
    }
    records: list[dict[str, Any]] = []
    for pic in drawing_root.findall(f".//{_XDR}pic"):
        blip = pic.find(f".//{_A}blip")
        embed = blip.get(f"{_R}embed") if blip is not None else None
        media_part = rel_targets.get(embed or "")
        media_bytes = file_map.get(media_part, b"") if media_part else b""
        cnvpr = pic.find(f"{_XDR}nvPicPr/{_XDR}cNvPr")
        if cnvpr is None:
            continue
        records.append(
            {
                "cnvpr": cnvpr,
                "anchor_key": _anchor_key_for_picture(pic),
                "media_part": media_part,
                "media_hash": _hash_bytes(media_bytes) if media_bytes else None,
            }
        )
    return records


def _match_picture_record(
    metadata: dict[str, Any],
    pictures: list[dict[str, Any]],
    used_indices: set[int],
) -> int | None:
    anchor_key = metadata.get("anchor_key")
    source_hash = metadata.get("source_hash")

    matchers = (
        lambda record: record.get("media_hash") == source_hash
        and record.get("anchor_key") == anchor_key,
        lambda record: record.get("media_hash") == source_hash,
        lambda record: anchor_key is not None and record.get("anchor_key") == anchor_key,
    )
    for matcher in matchers:
        candidates = [
            index
            for index, record in enumerate(pictures)
            if index not in used_indices and matcher(record)
        ]
        if len(candidates) == 1:
            return candidates[0]
    return None


def _anchor_key_for_picture(pic: Any) -> tuple[str, int, int] | None:
    parent = pic.getparent()
    if parent is None:
        return None
    from_el = parent.find(f"{_XDR}from")
    if from_el is None:
        return None
    col_el = from_el.find(f"{_XDR}col")
    row_el = from_el.find(f"{_XDR}row")
    if col_el is None or row_el is None:
        return None
    try:
        return ("cell", int(row_el.text or "0"), int(col_el.text or "0"))
    except ValueError:
        return None


def _normalize_anchor(anchor: Any) -> tuple[str, int, int] | None:
    if isinstance(anchor, str):
        row, col = coordinate_to_tuple(anchor)
        return ("cell", row - 1, col - 1)

    marker = getattr(anchor, "_from", None)
    if marker is not None and hasattr(marker, "row") and hasattr(marker, "col"):
        return ("cell", int(marker.row), int(marker.col))

    if hasattr(anchor, "row") and hasattr(anchor, "col"):
        return ("cell", int(anchor.row), int(anchor.col))

    return None


def _hash_file(path: Path) -> str:
    return _hash_bytes(path.read_bytes())


def _hash_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


# ─────────────────────────── CLI ──────────────────────────────────────────────

def _build_demo(out_path: Path) -> None:
    wb = create_workbook()
    ws = wb.active
    ws.title = "Demo"
    write_range(ws, 1, 1, [["Name", "Score", "Grade"], ["Alice", 95, "A"], ["Bob", 82, "B"]])
    format_cell(ws, 1, 1, font={"bold": True}, fill={"patternType": "solid", "fgColor": "4472C4"})
    format_cell(ws, 1, 2, font={"bold": True}, fill={"patternType": "solid", "fgColor": "4472C4"})
    format_cell(ws, 1, 3, font={"bold": True}, fill={"patternType": "solid", "fgColor": "4472C4"})
    ws2 = add_sheet(wb, "Formulas")
    for i in range(1, 6):
        write_cell(ws2, i, 1, i * 10)
    write_cell(ws2, 6, 1, "=SUM(A1:A5)")
    save(wb, out_path)
    print(f"Saved: {out_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Create a demo .xlsx file")
    parser.add_argument("--out", type=Path, default=Path("demo.xlsx"), help="Output path")
    args = parser.parse_args()
    _build_demo(args.out)


if __name__ == "__main__":
    main()
