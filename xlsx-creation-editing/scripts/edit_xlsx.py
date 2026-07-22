"""Edit existing Excel (.xlsx) workbooks.

Design decisions from primary sources:
- read_cell() loads with data_only=False so formula strings are preserved
  exactly as written; callers who need cached values should pass
  data_only=True to openpyxl.load_workbook().  openpyxl docs § Reading an
  existing file.
- read_all_values() iterates rows via ws.iter_rows(values_only=True) which
  respects merged cells and empty rows.
- replace_image_by_index() swaps the Image object in ws._images preserving
  the original anchor; openpyxl reads the new file on save.  This avoids
  touching the drawing XML relationship (openpyxl internal model).
- get_alt_text() accepts an lxml <xdr:pic> element and reads the descr
  attribute on <xdr:cNvPr> (ECMA-376 §20.5.2.8).
- iter_image_elements() opens the saved ZIP directly and parses each
  worksheet's drawing XML with lxml so callers get the raw element tree.

lxml boundary:
  Image alt text  — <xdr:cNvPr descr="…"> has no openpyxl setter
                    (ECMA-376 §20.5.2.8).  Reads and writes go through lxml.
"""

from __future__ import annotations

import argparse
import zipfile
from pathlib import Path
from typing import Any, Generator

import openpyxl
from lxml import etree
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

import create_xlsx

_XDR_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_XDR = f"{{{_XDR_NS}}}"
_REL_DRAWING = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"


# ─────────────────────────── cell / value reads ──────────────────────────────

def read_cell(ws: Any, row: int, col: int) -> Any:
    """Return the value at 1-based (*row*, *col*).

    Formula strings (starting with '=') are returned as-is when the workbook
    was opened with data_only=False (the default).
    """
    return ws.cell(row=row, column=col).value


def read_all_values(wb: Workbook) -> list[dict]:
    """Return a list of sheet dicts, each with 'sheet' and 'data' keys.

    'data' is a list of lists (rows × cols) of cell values.  Empty trailing
    rows are excluded.  openpyxl docs § Accessing many cells.
    """
    result = []
    for ws in wb.worksheets:
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        result.append({"sheet": ws.title, "data": rows})
    return result


# ─────────────────────────── image ops ───────────────────────────────────────

def get_alt_text(img_element: Any) -> str:
    """Read the alt text from an lxml <xdr:pic> element.

    ECMA-376 §20.5.2.8: <xdr:cNvPr descr="…"> inside <xdr:nvPicPr> carries
    the accessibility description for a picture in a SpreadsheetML drawing part.
    openpyxl has no high-level getter for this attribute, so we read it directly
    from the lxml element tree.
    """
    cnvpr = img_element.find(f"{_XDR}nvPicPr/{_XDR}cNvPr")
    if cnvpr is not None:
        return cnvpr.get("descr", "")
    return ""


def iter_image_elements(
    path: str | Path, ws_index: int = 0
) -> Generator[Any, None, None]:
    """Yield lxml <xdr:pic> elements from the drawing XML of a worksheet.

    Opens the saved ZIP directly.  Use get_alt_text() on each yielded element
    to read the alt text.

    ECMA-376 §20.5.2.8 — <xdr:pic> elements live inside anchors in the
    worksheet drawing part (xl/drawings/drawing{n}.xml).
    """
    path = Path(path)
    sheet_num = ws_index + 1
    with zipfile.ZipFile(path, "r") as zf:
        names = set(zf.namelist())
        rels_key = f"xl/worksheets/_rels/sheet{sheet_num}.xml.rels"
        if rels_key not in names:
            return
        rels_root = etree.fromstring(zf.read(rels_key))
        drawing_target: str | None = None
        for rel in rels_root:
            if _REL_DRAWING in rel.get("Type", ""):
                drawing_target = rel.get("Target", "")
                break
        if not drawing_target:
            return
        drawing_part = _resolve_path(
            f"xl/worksheets/sheet{sheet_num}.xml", drawing_target
        )
        if drawing_part not in names:
            return
        root = etree.fromstring(zf.read(drawing_part))
        yield from root.findall(f".//{_XDR}pic")


def replace_image_by_index(
    ws: Any,
    index: int,
    new_path: str | Path,
    *,
    alt_text: str | None = None,
) -> bool:
    """Replace the *index*-th image on a worksheet with a new image file.

    Swaps the Image object in ws._images preserving the original anchor.
    If *alt_text* is provided, schedules it for injection on the next save()
    via the ws._xlsx_skill_alt_texts side-channel.

    Returns True on success, False if *index* is out of range.
    """
    if not (0 <= index < len(ws._images)):
        return False
    old_img = ws._images[index]
    new_img = XLImage(str(new_path))
    new_img.anchor = old_img.anchor
    ws._images[index] = new_img
    create_xlsx.set_image_metadata(
        ws,
        index,
        new_path,
        old_img.anchor,
        alt_text=alt_text,
    )
    return True


def _resolve_path(source: str, target: str) -> str:
    # Absolute OPC target paths start with '/' — strip the leading slash.
    if target.startswith("/"):
        return target.lstrip("/")
    source_dir = source.rsplit("/", 1)[0] if "/" in source else ""
    combined = source_dir + "/" + target if source_dir else target
    parts = combined.split("/")
    resolved: list[str] = []
    for part in parts:
        if part == "..":
            if resolved:
                resolved.pop()
        elif part and part != ".":
            resolved.append(part)
    return "/".join(resolved)


# ─────────────────────────── CLI ──────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Read values from an .xlsx file")
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_read = sub.add_parser("read", help="Print all cell values")
    p_read.add_argument("file", type=Path)

    p_cell = sub.add_parser("cell", help="Read a single cell value")
    p_cell.add_argument("file", type=Path)
    p_cell.add_argument("sheet", type=str, help="Sheet name")
    p_cell.add_argument("row", type=int)
    p_cell.add_argument("col", type=int)

    args = parser.parse_args()

    if args.cmd == "read":
        wb = openpyxl.load_workbook(str(args.file))
        for entry in read_all_values(wb):
            print(f"\n=== {entry['sheet']} ===")
            for row in entry["data"]:
                print(row)

    elif args.cmd == "cell":
        wb = openpyxl.load_workbook(str(args.file))
        ws = wb[args.sheet]
        print(read_cell(ws, args.row, args.col))


if __name__ == "__main__":
    main()
