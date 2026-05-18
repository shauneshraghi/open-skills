"""Comprehensive test suite against the Apache POI XLSX corpus.

Covers:
- Files with multiple sheets, merged cells, conditional formatting
- Date format handling, boolean values, inline strings
- Pivot tables, named ranges, data validation
- Sheet reorder stability
- Cell value idempotency
- Image listing (no crash)
"""

from __future__ import annotations

import os
import struct
import sys
import tempfile
import traceback
import zlib
from pathlib import Path
from typing import NamedTuple

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl

from scripts.create_xlsx import (
    add_image_to_sheet,
    create_workbook,
    merge_cells,
    reorder_sheet,
    save,
    write_cell,
)
from scripts.edit_xlsx import read_all_values
from scripts.validate_xlsx import validate_xlsx

POI_PATH = Path(
    os.environ.get(
        "POI_PATH",
        str(Path(__file__).resolve().parents[3] / "poi" / "test-data" / "spreadsheet"),
    )
)

CORPUS_FIXTURES = [
    "sample.xlsx",
    "simple-monthly-budget.xlsx",
    "ExcelTables.xlsx",
    "Formatting.xlsx",
    "ConditionalFormattingSamples.xlsx",
    "123233_charts.xlsx",
    "styles.xlsx",
    "DateFormatTests.xlsx",
    "AverageTaxRates.xlsx",
    "simple-table-named-range.xlsx",
    "reordered_sheets.xlsx",
    "Booleans.xlsx",
    "InlineStrings.xlsx",
    "DataValidations-49244.xlsx",
    "ExcelPivotTableSample.xlsx",
]


class TestResult(NamedTuple):
    fixture: str
    test_name: str
    passed: bool
    detail: str = ""


results: list[TestResult] = []


def record(fixture: str, test_name: str, passed: bool, detail: str = "") -> None:
    results.append(TestResult(fixture, test_name, passed, detail))
    status = "PASS" if passed else "FAIL"
    print(f"  [{status}] {test_name}: {detail}" if detail else f"  [{status}] {test_name}")


# ─────────────────────────── check 1 ─────────────────────────────────────────

def check_open_and_iterate(path: Path) -> None:
    fname = path.name
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
        non_none = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                non_none += sum(1 for v in row if v is not None)
        record(fname, "open_and_iterate", True, f"{len(wb.sheetnames)} sheets, {non_none} non-None cells")
    except Exception as exc:
        record(fname, "open_and_iterate", False, str(exc))


# ─────────────────────────── check 2 ─────────────────────────────────────────

def check_read_all_values_structure(path: Path) -> None:
    fname = path.name
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
        result = read_all_values(wb)
        ok = (
            isinstance(result, list)
            and len(result) == len(wb.sheetnames)
            and all("sheet" in e and "data" in e for e in result)
            and all(isinstance(e["data"], list) for e in result)
            and all(
                all(isinstance(row, list) for row in e["data"])
                for e in result
            )
        )
        record(fname, "read_all_values_structure", ok, f"{len(result)} entries")
    except Exception as exc:
        record(fname, "read_all_values_structure", False, str(exc))


# ─────────────────────────── check 3 ─────────────────────────────────────────

def check_sheet_count_positive(path: Path) -> None:
    fname = path.name
    try:
        wb = openpyxl.load_workbook(str(path))
        ok = len(wb.sheetnames) >= 1
        record(fname, "sheet_count_positive", ok, f"{len(wb.sheetnames)} sheets")
    except Exception as exc:
        record(fname, "sheet_count_positive", False, str(exc))


# ─────────────────────────── check 4 ─────────────────────────────────────────

def check_validate_xlsx(path: Path) -> None:
    fname = path.name
    try:
        vr = validate_xlsx(path)
        record(
            fname, "validate_xlsx", vr.valid,
            f"sheets={vr.info.get('sheet_count', '?')}"
            + (f", errors={vr.errors[:1]}" if not vr.valid else ""),
        )
    except Exception as exc:
        record(fname, "validate_xlsx", False, str(exc))


# ─────────────────────────── check 5 ─────────────────────────────────────────

def check_write_cell_idempotency(path: Path) -> None:
    fname = path.name
    tmpdir = Path(tempfile.mkdtemp())
    out = tmpdir / path.name
    sentinel = "xlsx_skill_idempotency_check"
    try:
        wb = openpyxl.load_workbook(str(path))
        ws = wb.worksheets[0]
        write_cell(ws, 1, 1, sentinel)
        save(wb, out)
        wb2 = openpyxl.load_workbook(str(out))
        actual = wb2.worksheets[0].cell(row=1, column=1).value
        record(fname, "write_cell_idempotency", actual == sentinel, f"got {actual!r}")
    except Exception as exc:
        record(fname, "write_cell_idempotency", False, str(exc))


# ─────────────────────────── check 6 ─────────────────────────────────────────

def check_reorder_stability(path: Path) -> None:
    fname = path.name
    tmpdir = Path(tempfile.mkdtemp())
    out = tmpdir / path.name
    try:
        wb = openpyxl.load_workbook(str(path))
        n = len(wb.worksheets)
        if n < 2:
            record(fname, "reorder_stability", True, "skipped (1 sheet)")
            return
        original = list(wb.sheetnames)
        reorder_sheet(wb, n - 1, 0)
        reorder_sheet(wb, 0, n - 1)
        save(wb, out)
        wb2 = openpyxl.load_workbook(str(out))
        match = list(wb2.sheetnames) == original
        record(fname, "reorder_stability", match, f"{'restored' if match else 'MISMATCH'} ({n} sheets)")
    except Exception as exc:
        record(fname, "reorder_stability", False, str(exc))


# ─────────────────────────── check 7 ─────────────────────────────────────────

def check_merge_cells_roundtrip(path: Path) -> None:
    fname = path.name
    tmpdir = Path(tempfile.mkdtemp())
    out = tmpdir / path.name
    try:
        wb = openpyxl.load_workbook(str(path))
        ws = wb.worksheets[0]
        # Unmerge any ranges that overlap with B2:C3 to avoid conflicts.
        overlapping = [
            str(mr) for mr in list(ws.merged_cells.ranges)
            if _overlaps_b2c3(str(mr))
        ]
        for rng in overlapping:
            try:
                ws.unmerge_cells(rng)
            except Exception:
                pass
        merge_cells(ws, 2, 2, 3, 3)
        save(wb, out)
        wb2 = openpyxl.load_workbook(str(out))
        ws2 = wb2.worksheets[0]
        merged_ranges = str(ws2.merged_cells)
        ok = "B2:C3" in merged_ranges
        record(fname, "merge_cells_roundtrip", ok, f"B2:C3 {'found' if ok else 'NOT found'} in {merged_ranges!r}")
    except Exception as exc:
        record(fname, "merge_cells_roundtrip", False, str(exc))


def _overlaps_b2c3(rng: str) -> bool:
    """Return True if a merge range string overlaps B2:C3 (cols 2-3, rows 2-3)."""
    try:
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(rng)
        return not (max_col < 2 or min_col > 3 or max_row < 2 or min_row > 3)
    except Exception:
        return False


# ─────────────────────────── check 8 ─────────────────────────────────────────

def check_image_list(path: Path) -> None:
    fname = path.name
    try:
        wb = openpyxl.load_workbook(str(path))
        total = sum(len(ws._images) for ws in wb.worksheets)
        record(fname, "image_list", True, f"{total} image(s) found")
    except Exception as exc:
        record(fname, "image_list", False, str(exc))


# ─────────────────────────── main ────────────────────────────────────────────

def main() -> None:
    if not POI_PATH.exists():
        print(f"ERROR: POI_PATH not found: {POI_PATH}")
        sys.exit(1)

    print(f"POI corpus: {POI_PATH}")
    print("=" * 70)

    for fname in CORPUS_FIXTURES:
        fixture = POI_PATH / fname
        if not fixture.exists():
            print(f"\n[SKIP] {fname} not found")
            continue
        print(f"\n--- {fname} ---")
        check_open_and_iterate(fixture)
        check_read_all_values_structure(fixture)
        check_sheet_count_positive(fixture)
        check_validate_xlsx(fixture)
        check_write_cell_idempotency(fixture)
        check_reorder_stability(fixture)
        check_merge_cells_roundtrip(fixture)
        check_image_list(fixture)

    print("\n" + "=" * 70)
    passed = sum(1 for r in results if r.passed)
    failed = sum(1 for r in results if not r.passed)
    print(f"Corpus test results: {passed} passed, {failed} failed (total {len(results)})")

    if failed:
        print("\nFailures:")
        for r in results:
            if not r.passed:
                print(f"  FAIL {r.fixture} / {r.test_name}: {r.detail}")
        sys.exit(1)
    else:
        sys.exit(0)


if __name__ == "__main__":
    main()
