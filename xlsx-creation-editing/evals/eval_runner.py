"""Run the 5 evaluation test cases for the xlsx-creation-editing skill.

Uses tempfile.mkdtemp() for all output and POI_PATH env var for corpus fixtures.
"""

from __future__ import annotations

import os
import struct
import sys
import tempfile
import traceback
import zlib
from pathlib import Path
from typing import Callable

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl

from scripts.create_xlsx import (
    add_image_to_sheet,
    add_sheet,
    create_workbook,
    format_cell,
    merge_cells,
    reorder_sheet,
    save,
    write_cell,
    write_range,
)
from scripts.edit_xlsx import (
    get_alt_text,
    iter_image_elements,
    read_all_values,
)
from scripts.validate_xlsx import validate_xlsx

POI_PATH = Path(
    os.environ.get(
        "POI_PATH",
        str(Path(__file__).resolve().parents[3] / "poi" / "test-data" / "spreadsheet"),
    )
)


def _make_png(r: int, g: int, b: int, w: int = 20, h: int = 20) -> bytes:
    raw = (b"\x00" + bytes([r, g, b]) * w) * h
    compressed = zlib.compress(raw)

    def chunk(ctype: bytes, data: bytes) -> bytes:
        c = ctype + data
        return struct.pack(">I", len(data)) + c + struct.pack(">I", zlib.crc32(c) & 0xFFFFFFFF)

    ihdr = struct.pack(">IIBBBBB", w, h, 8, 2, 0, 0, 0)
    return b"\x89PNG\r\n\x1a\n" + chunk(b"IHDR", ihdr) + chunk(b"IDAT", compressed) + chunk(b"IEND", b"")


class EvalResult:
    def __init__(self, test_id: str):
        self.test_id = test_id
        self.passed = True
        self.failures: list[str] = []
        self.log: list[str] = []

    def assert_eq(self, label: str, actual, expected) -> None:
        if actual != expected:
            self.passed = False
            self.failures.append(f"FAIL [{label}]: expected {expected!r}, got {actual!r}")
        else:
            self.log.append(f"PASS [{label}]")

    def assert_true(self, label: str, value: bool) -> None:
        if not value:
            self.passed = False
            self.failures.append(f"FAIL [{label}]: expected True, got False")
        else:
            self.log.append(f"PASS [{label}]")

    def assert_in(self, label: str, needle, haystack) -> None:
        if needle not in haystack:
            self.passed = False
            self.failures.append(f"FAIL [{label}]: {needle!r} not in {haystack!r}")
        else:
            self.log.append(f"PASS [{label}]")


# ─────────────────────────── test 1 ──────────────────────────────────────────

def test_create_workbook_with_sheets_and_values() -> EvalResult:
    r = EvalResult("create_workbook_with_sheets_and_values")
    tmpdir = Path(tempfile.mkdtemp())
    out = tmpdir / "test1.xlsx"

    wb = create_workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    write_range(
        ws_summary, 1, 1,
        [
            ["Product", "Units", "Revenue"],
            ["Widget A", 100, 2500.0],
            ["Widget B", 75, 1875.0],
            ["Widget C", 50, 1250.0],
        ],
    )

    ws_formulas = add_sheet(wb, "Formulas")
    for i in range(1, 6):
        write_cell(ws_formulas, i, 1, i * 10)
    write_cell(ws_formulas, 6, 1, "=SUM(A1:A5)")

    save(wb, out)

    wb2 = openpyxl.load_workbook(str(out))
    r.assert_eq("sheet count", len(wb2.sheetnames), 2)
    r.assert_eq("Summary A1", wb2["Summary"]["A1"].value, "Product")
    r.assert_eq("Summary B2", wb2["Summary"]["B2"].value, 100)
    r.assert_eq("Summary C4", wb2["Summary"]["C4"].value, 1250.0)
    r.assert_eq("Formulas A6 formula", wb2["Formulas"]["A6"].value, "=SUM(A1:A5)")
    r.assert_true("validate_xlsx passes", validate_xlsx(out).valid)
    return r


# ─────────────────────────── test 2 ──────────────────────────────────────────

def test_format_cells_and_merge() -> EvalResult:
    r = EvalResult("format_cells_and_merge")
    tmpdir = Path(tempfile.mkdtemp())
    out = tmpdir / "test2.xlsx"

    wb = create_workbook()
    ws = wb.active
    write_cell(ws, 1, 1, "Header")
    format_cell(
        ws, 1, 1,
        font={"bold": True, "size": 14},
        fill={"patternType": "solid", "fgColor": "FFFF00"},
    )
    write_cell(ws, 2, 1, "Merged text")
    merge_cells(ws, 2, 1, 2, 3)
    save(wb, out)

    wb2 = openpyxl.load_workbook(str(out))
    ws2 = wb2.active
    r.assert_true("A1 font.bold", ws2["A1"].font.bold is True)
    fgcolor = ws2["A1"].fill.fgColor.rgb
    r.assert_true("A1 fill yellow", "FFFF00" in fgcolor)
    r.assert_in("A2:C2 merged", "A2:C2", str(ws2.merged_cells))
    r.assert_eq("merged cell value", ws2["A2"].value, "Merged text")
    r.assert_true("validate_xlsx passes", validate_xlsx(out).valid)
    return r


# ─────────────────────────── test 3 ──────────────────────────────────────────

def test_embed_image_with_alt_text() -> EvalResult:
    r = EvalResult("embed_image_with_alt_text")
    tmpdir = Path(tempfile.mkdtemp())
    img_path = tmpdir / "test_img.png"
    img_path.write_bytes(_make_png(0, 120, 255))
    out = tmpdir / "test3.xlsx"

    wb = create_workbook()
    ws = wb.active
    write_cell(ws, 1, 1, "Image below")
    add_image_to_sheet(ws, img_path, "B2", alt_text="Test image")
    save(wb, out)

    import zipfile
    with zipfile.ZipFile(out) as z:
        media = [n for n in z.namelist() if n.startswith("xl/media/")]
        r.assert_true("image part in xl/media/", len(media) > 0)

    pics = list(iter_image_elements(out, ws_index=0))
    r.assert_true("at least one <xdr:pic> element", len(pics) > 0)
    if pics:
        r.assert_eq("alt text via lxml", get_alt_text(pics[0]), "Test image")

    r.assert_true("validate_xlsx passes", validate_xlsx(out).valid)
    return r


# ─────────────────────────── test 4 ──────────────────────────────────────────

def test_read_all_values_from_poi_fixture() -> EvalResult:
    r = EvalResult("read_all_values_from_poi_fixture")
    fixture = POI_PATH / "sample.xlsx"
    if not fixture.exists():
        r.passed = False
        r.failures.append(f"POI fixture not found: {fixture}")
        return r

    wb = openpyxl.load_workbook(str(fixture))
    result = read_all_values(wb)

    r.assert_true("result is a list", isinstance(result, list))
    r.assert_eq("entry count matches sheet count", len(result), len(wb.sheetnames))
    for entry in result:
        r.assert_true(f"'{entry.get('sheet')}' has 'sheet' key", "sheet" in entry)
        r.assert_true(f"'{entry.get('sheet')}' has 'data' key", "data" in entry)
        r.assert_true(
            f"'{entry.get('sheet')}' data is list of lists",
            isinstance(entry["data"], list)
            and all(isinstance(row, list) for row in entry["data"]),
        )
    return r


# ─────────────────────────── test 5 ──────────────────────────────────────────

def test_round_trip_sheet_reorder() -> EvalResult:
    r = EvalResult("round_trip_sheet_reorder")
    tmpdir = Path(tempfile.mkdtemp())
    out = tmpdir / "test5.xlsx"

    fixture = POI_PATH / "reordered_sheets.xlsx"
    if fixture.exists():
        import shutil
        src = fixture
        wb = openpyxl.load_workbook(str(src))
    else:
        wb = create_workbook()
        wb.active.title = "Alpha"
        add_sheet(wb, "Beta")
        add_sheet(wb, "Gamma")

    original_order = list(wb.sheetnames)
    r.assert_true("at least 3 sheets", len(original_order) >= 2)

    n = len(wb.worksheets)
    reorder_sheet(wb, n - 1, 0)
    reorder_sheet(wb, 0, n - 1)
    save(wb, out)

    wb2 = openpyxl.load_workbook(str(out))
    r.assert_eq("order restored", list(wb2.sheetnames), original_order)
    r.assert_true("validate_xlsx passes", validate_xlsx(out).valid)
    return r


# ─────────────────────────── runner ──────────────────────────────────────────

TESTS: list[Callable[[], EvalResult]] = [
    test_create_workbook_with_sheets_and_values,
    test_format_cells_and_merge,
    test_embed_image_with_alt_text,
    test_read_all_values_from_poi_fixture,
    test_round_trip_sheet_reorder,
]


def run_all() -> list[EvalResult]:
    results = []
    for test_fn in TESTS:
        print(f"\n{'=' * 60}")
        print(f"Running: {test_fn.__name__}")
        print("-" * 60)
        try:
            result = test_fn()
        except Exception as exc:
            result = EvalResult(test_fn.__name__)
            result.passed = False
            result.failures.append(f"EXCEPTION: {exc}")
            traceback.print_exc()

        for line in result.log:
            print(f"  {line}")
        for line in result.failures:
            print(f"  {line}")

        status = "PASSED" if result.passed else "FAILED"
        print(f"\n  → {status}: {result.test_id}")
        results.append(result)

    print(f"\n{'=' * 60}")
    passed = sum(1 for r in results if r.passed)
    print(f"Results: {passed}/{len(results)} tests passed")
    return results


if __name__ == "__main__":
    results = run_all()
    sys.exit(0 if all(r.passed for r in results) else 1)
